"""Extract text/images from user-imported attachments for one chat request."""
import base64
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_TOTAL_CHARS = 120_000
MAX_ATTACHMENTS = 6
MAX_IMAGES = 4

PLAIN_TEXT_SUFFIXES = {
    ".txt", ".md", ".markdown", ".csv", ".tsv", ".json", ".xml", ".log",
    ".html", ".htm", ".ini", ".cfg", ".yaml", ".yml", ".toml", ".py",
    ".js", ".ts", ".css", ".sql", ".sh", ".bat",
}
DOCUMENT_SUFFIXES = PLAIN_TEXT_SUFFIXES | {".pdf", ".doc", ".docx", ".xlsx", ".xls"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
SUPPORTED = DOCUMENT_SUFFIXES | IMAGE_SUFFIXES
IMAGE_MEDIA_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
    ".bmp": "image/bmp",
}


def _read_text(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _read_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("读取 PDF 需要安装 pypdf") from exc
    reader = PdfReader(str(path))
    pages = [(page.extract_text() or "") for page in reader.pages]
    return "\n\n".join(pages)


def _read_doc(path: Path) -> str:
    # .doc 是旧式二进制格式，Windows 上优先使用已安装的 Microsoft Word。
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError("读取 .doc 需要 Windows Microsoft Word 或 pywin32") from exc

    pythoncom.CoInitialize()
    word = None
    document = None
    try:
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        word.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
        document = word.Documents.Open(str(path), ReadOnly=True, AddToRecentFiles=False)
        return document.Content.Text
    finally:
        if document is not None:
            try:
                document.Close(False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _read_docx(path: Path) -> str:
    # .docx 本质是 zip 包，直接解析 word/document.xml，无需额外依赖。
    import zipfile
    import xml.etree.ElementTree as ElementTree

    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    try:
        with zipfile.ZipFile(str(path)) as archive:
            with archive.open("word/document.xml") as handle:
                tree = ElementTree.parse(handle)
    except (KeyError, zipfile.BadZipFile) as exc:
        raise RuntimeError("DOCX 文件已损坏或不是有效的 Office 文档") from exc

    paragraphs = []
    for paragraph in tree.getroot().iter("%sp" % namespace):
        texts = [node.text or "" for node in paragraph.iter("%st" % namespace)]
        paragraphs.append("".join(texts))
    return "\n".join(paragraphs)


def _read_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 XLSX 需要安装 openpyxl") from exc

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    sections = []
    try:
        for sheet in workbook.worksheets:
            rows = _format_sheet_rows(sheet.iter_rows(values_only=True))
            if rows:
                sections.append("工作表：%s\n%s" % (sheet.title, "\n".join(rows)))
    finally:
        try:
            workbook.close()
        except Exception:
            pass
    return "\n\n".join(sections)


def _read_xls(path: Path) -> str:
    # .xls 是旧式二进制格式，Windows 上优先使用已安装的 Microsoft Excel。
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError("读取 .xls 需要 Windows Microsoft Excel 或 pywin32") from exc

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AutomationSecurity = 3
        workbook = excel.Workbooks.Open(str(path), ReadOnly=True, UpdateLinks=0)
        sections = []
        for sheet in workbook.Worksheets:
            values = sheet.UsedRange.Value
            if values is None:
                continue
            if not isinstance(values, tuple):
                values = ((values,),)
            rows = []
            for row in values:
                if not isinstance(row, tuple):
                    row = (row,)
                rows.append(row)
            formatted = _format_sheet_rows(rows)
            if formatted:
                sections.append("工作表：%s\n%s" % (sheet.Name, "\n".join(formatted)))
        return "\n\n".join(sections)
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _format_sheet_rows(rows: Iterable[Iterable]) -> List[str]:
    formatted = []
    for row in rows or ():
        cells = ["" if value is None else str(value).strip() for value in row]
        if any(cells):
            formatted.append("\t".join(cells).rstrip())
    return formatted


def _read_document(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _read_pdf(path)
    if suffix == ".doc":
        return _read_doc(path)
    if suffix == ".docx":
        return _read_docx(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".xls":
        return _read_xls(path)
    return _read_text(path)


def _resolve(path_string: str) -> Path:
    path = Path(path_string).expanduser().resolve()
    if not path.is_file():
        raise RuntimeError("文件不存在")
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED:
        raise RuntimeError("不支持的文件类型 %s" % (suffix or "(无扩展名)"))
    if path.stat().st_size > MAX_FILE_BYTES:
        raise RuntimeError("单个文件不能超过 10 MB")
    return path


def extract_content(path_string: str) -> Dict[str, str]:
    """Return ``{"kind": "image", ...}`` or ``{"kind": "text", ...}`` for one file."""
    path = _resolve(path_string)
    suffix = path.suffix.lower()
    if suffix in IMAGE_SUFFIXES:
        return {
            "kind": "image",
            "name": path.name,
            "media_type": IMAGE_MEDIA_TYPES[suffix],
            "path": str(path),
        }
    text = _read_document(path).strip()
    if not text:
        if suffix == ".pdf":
            raise RuntimeError("未提取到文字；扫描版 PDF 暂不支持 OCR")
        raise RuntimeError("文档中没有可读取的文字")
    return {"kind": "text", "name": path.name, "text": text, "path": str(path)}


def read_image_data(path_string: str) -> Optional[Dict[str, str]]:
    """Base64-encode a stored image at request time; return None when unusable."""
    try:
        path = Path(path_string).expanduser().resolve()
        suffix = path.suffix.lower()
        if suffix not in IMAGE_SUFFIXES or not path.is_file():
            return None
        if path.stat().st_size > MAX_FILE_BYTES:
            return None
        return {
            "media_type": IMAGE_MEDIA_TYPES[suffix],
            "data": base64.b64encode(path.read_bytes()).decode("ascii"),
        }
    except OSError:
        return None


def load_attachments(attachments: Iterable[dict]) -> Tuple[List[dict], List[dict], List[str]]:
    """Split attachments into ``(documents, images, errors)`` for one request."""
    documents: List[dict] = []
    images: List[dict] = []
    errors: List[str] = []
    total_chars = 0
    for attachment in list(attachments or [])[:MAX_ATTACHMENTS]:
        if not isinstance(attachment, dict):
            continue
        try:
            item = extract_content(str(attachment.get("path", "")))
        except Exception as exc:
            errors.append("%s：%s" % (attachment.get("name", "文件"), exc))
            continue
        if item["kind"] == "image":
            if len(images) >= MAX_IMAGES:
                errors.append("%s：图片附件最多 %d 张" % (item["name"], MAX_IMAGES))
                continue
            images.append({
                "name": item["name"],
                "media_type": item["media_type"],
                "path": item["path"],
                "url": attachment.get("url", ""),
            })
            continue
        remaining = MAX_TOTAL_CHARS - total_chars
        if remaining <= 0:
            errors.append("附件文本总量超过限制")
            break
        text = item["text"][:remaining]
        documents.append({"name": item["name"], "text": text, "path": item["path"]})
        total_chars += len(text)
    return documents, images, errors
